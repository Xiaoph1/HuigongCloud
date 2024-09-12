import pandas as pd
import openpyxl
import json

class ExcelToJsonConverter:

    __inputpath = ""
    __outputpath = ""
    __orders = {}
    __workline = {}
    __sourcetime = {}
    __lockedwork = {}
    __machinechange = {}
    __scheduling_strategy = []

    def convert_scheduling_strategy(self):
        try:
            df = pd.read_excel(self.input_path, sheet_name="排程策略")
            self.__scheduling_strategy = df.to_dict('records')
        except ValueError:
            print("排程策略工作簿不存在，跳过。")

    def __init__(self, input_path, output_path):
        self.input_path = input_path
        self.output_path = output_path

    def convert(self):
        df = pd.read_excel(self.input_path, sheet_name = "工单")

        global counts
        # 解析工单数据
        orders = {}
        for index, row in df.iterrows():
            order_id = row['工单编号']
            orders[order_id] = {
                '工单编号': row['工单编号'],
                '物料编码': row['物料编号'],
                '数量': row['数量'],
                '前置工单': "/" if pd.isna(row['前置工单']) else row['前置工单'],
                '排程策略': row['排程策略'],
                '计划开始时间': row['计划开始日期'],
                '计划完工时间': row['计划完工日期']
            }
        self.__orders = orders
        
    #解析工艺路线数据
    def convert_workline(self):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb["工艺路线"]

        workline = {}

        current_mo, current_fg, current_process = None, None, None
        resources = []  # 用于收集当前工序的所有资源

        for row in ws.iter_rows(min_row=3, values_only=True):
            mo_id, fg_id, process, pre_process, resource, resource_count, prep_time, work_time, post_time, cross_shift = row

            # 检测到新的工单编号和物料编码
            if mo_id and fg_id:
                if current_process:  # 如果之前的工序有未保存的资源，先保存
                    workline[current_mo][current_fg]["资源需求"].append(resources)
                resources = []  # 重置资源列表
                current_mo, current_fg = mo_id, fg_id
                if current_mo not in workline:
                    workline[current_mo] = {}
                workline[current_mo][current_fg] = {"工序列表": [], "前置工序": [], "资源需求": [], "准备工时": [], "作业工时": [], "后置工时": [], "是否可跨班组": []}
                pass_process = current_process
                current_process = None  # 重置当前工序

            # 检测到新的工序
            if process:
                if current_process:
                    workline[current_mo][current_fg]["资源需求"].append(resources)
                resources = []  # 重置资源列表
                current_process = process
                workline[current_mo][current_fg]["工序列表"].append(process)
                workline[current_mo][current_fg]["前置工序"].append(pre_process if pre_process else "/")
                workline[current_mo][current_fg]["准备工时"].append(int(prep_time))
                workline[current_mo][current_fg]["作业工时"].append(int(work_time))
                workline[current_mo][current_fg]["后置工时"].append(int(post_time))
                workline[current_mo][current_fg]["是否可跨班组"].append(cross_shift)

            # 收集资源信息
            if resource and resource_count and current_process:  # 确保在有工序的情况下收集资源
                resources.append(f"{resource}{resource_count}")
        
        # 循环结束后，保存最后一个工序的资源
        if current_process:
            workline[current_mo][current_fg]["资源需求"].append(resources)
        
        self.__workline = workline

    #解析资源日历数据
    def convert_sourcetime(self):
        df = pd.read_excel(self.input_path, sheet_name="资源日历")

    # 解析数据前先过滤掉脏数据，假设脏数据的标识是资源编号为空
        df = df[df['资源编号'].notna()]

        sourcetime = {}
        for index, row in df.iterrows():
            order_id = row['资源编号']
            sourcetime[order_id] = {
                '资源编号': row['资源编号'],
                '资源名称': row['资源名称'],
                '资源分类': row['资源分类'],
                '资源数量': row['资源数量'],
                '开始日期': row['开始日期'],
                '结束日期': row['结束日期'],
                '开始时间': row['开始时间'],
                '结束时间': row['结束时间'],
                '优先级': row['优先级']
            }
        self.__sourcetime = sourcetime

    #解析锁排程数据
    def convert_lockedwork(self):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb["锁排程"]

        lockedwork = {}

        current_mo, current_fg, current_process = None, None, None
        resources = []  #用于收集当前工序的所有资源
        resources_id = []
        resources_count = []
        resources_begintime = []
        resources_endtime = []

        for row in ws.iter_rows(min_row = 2, values_only=True):
            mo_id, fg_id, process, resource, resource_id, resource_count, begin_time, end_time = row

            #检测到新的工单编号或物料编号
            if mo_id and fg_id:
                if current_process:  # 如果之前的工序有未保存的资源，先保存
                    lockedwork[current_mo][current_fg]["资源池"].append(resources)
                    lockedwork[current_mo][current_fg]['资源ID'].append(resources_id)
                    lockedwork[current_mo][current_fg]['资源需求'].append(resources_count)
                    lockedwork[current_mo][current_fg]['开始时间'].append(resources_begintime)
                    lockedwork[current_mo][current_fg]['结束时间'].append(resources_endtime)
                resources = [] #重置资源列表
                resources_id = []
                resources_count = []
                resources_begintime = []
                resources_endtime = []
                current_mo, current_fg = mo_id, fg_id
                if current_mo not in lockedwork:
                    lockedwork[current_mo] = {}
                lockedwork[current_mo][current_fg] = {"工序": process, "资源池": [], "资源ID": [], "资源需求": [], "开始时间": [], "结束时间": []}
                current_process = None #重置当前工序
            
            #检测到新的工序
            if process:
                if current_process:
                    lockedwork[current_mo][current_fg]["资源池"].append(resources)
                    lockedwork[current_mo][current_fg]['资源ID'].append(resources_id)
                    lockedwork[current_mo][current_fg]['资源需求'].append(resources_count)
                    lockedwork[current_mo][current_fg]['开始时间'].append(resources_begintime)
                    lockedwork[current_mo][current_fg]['结束时间'].append(resources_endtime)
                resources = []
                resources_id = []
                resources_count = []
                resources_begintime = []
                resources_endtime = []
                current_process = process

            # 收集资源信息
            if resource and resource_count and current_process:
                resources.append(resource)
                resources_id.append(resource_id)
                resources_count.append(resource_count)
                resources_begintime.append(begin_time)
                resources_endtime.append(end_time)
        if current_process:
            lockedwork[current_mo][current_fg]["资源池"].append(resources)
            lockedwork[current_mo][current_fg]['资源ID'].append(resources_id)
            lockedwork[current_mo][current_fg]['资源需求'].append(resources_count)
            lockedwork[current_mo][current_fg]['开始时间'].append(resources_begintime)
            lockedwork[current_mo][current_fg]['结束时间'].append(resources_endtime)
        
        self.__lockedwork = lockedwork
    
    def convert_machinechange(self):
        df = pd.read_excel(self.input_path, sheet_name = '设备换型', index_col = 0)
        
        index = df.index.tolist()

        machinechange = {
            '工序列表': index,
            '换型矩阵': []
        }
        change = []
        temp = []
        for index1, row in df.iterrows():
            for i in range(0, len(index)):
                temp.append(str(row[index[i]]))
            machinechange['换型矩阵'].append(temp)
            temp = []
        
        self.__machinechange = machinechange

    def write_json(self):
        # 构建JSON结构
        data = {
            '全部工单': self.__orders,
            '工艺路线': self.__workline,
            '资源日历': self.__sourcetime,
            '锁排程': self.__lockedwork,
            '设备换型': self.__machinechange,
            '排程策略': self.__scheduling_strategy
        }

        # 转换为JSON
        json_str = json.dumps(data, ensure_ascii=False, indent=4)

        # 保存JSON文件
        with open(self.output_path, 'w', encoding='utf-8') as f:
            f.write(json_str)

# 使用示例
if __name__ == '__main__':
   
    converter = ExcelToJsonConverter("ss_3.xlsx", "output.json")
    converter.convert()
    converter.convert_workline()
    converter.convert_sourcetime()
    converter.convert_lockedwork()
    converter.convert_machinechange()
    converter.convert_scheduling_strategy()
    converter.write_json()